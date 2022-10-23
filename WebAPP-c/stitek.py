from PIL import Image, ImageFont, ImageDraw
from openpyxl import load_workbook
import qrcode
import csv
import datetime
from datetime import date

def stitek():
    data = []

    with open("data/data.csv") as file:
        reader = csv.DictReader(file)

        for r in reader:
            data.append({'QR': r['QR'],'Firma': r['Firma'],'SF': r['SF'],'DBZ': r['DBZ'],'WB': r['WB'],'MLFB': r['MLFB'],'Z': r['Z'],'cislo': r['cislo'],'Datum': r['Datum']})

    #print((data[0])["MLFB"])
    QR = (data[0])['QR']
    Firma = (data[0])['Firma']
    SF = (data[0])['SF']
    DBZ = (data[0])['DBZ']
    WB = (data[0])['WB']
    MLFB = (data[0])['MLFB']
    Z = (data[0])['Z']
    cislo = (data[0])['cislo']
    Datum = (data[0])['Datum']

    if Z == "/" :
        Z = ""
        
    else:
        Z = f"Z:{Z}"
        
    print("OK")
    return(QR,Firma,SF,DBZ,WB,MLFB,Z,cislo,Datum)

def vyroba(QR):
    wb = load_workbook(filename = 'data/data.xlsx')
    ws = wb['nic']
    
    imput = QR[2:50]
    cd = len(imput)
    #print(cd)
    
#    if " " in QR :
#        return render_template('error/error_qr.html',QR=QR)        

    if cd >= 37 :
        MLFB = imput[0:20]
        Z = "?"
        barvaZ = "#ff0000"
        #print(z)
        #print(MLFB)
        datum = imput[24:26]
        #print(datum)
    
        vc = imput[24:38]
        wc = len(vc)
        #print(wc)


        if wc == 13 :

            cislo = imput[24:28]+"-"+imput[28:32]+"-"+imput[32:34]+"-"+imput[34:37]
            #print(cislo)

        elif wc == 14:     

            cislo = imput[24:29]+"-"+imput[29:33]+"-"+imput[33:35]+"-"+imput[35:38]
            #print(cislo)
    
    elif cd < 37 :
        MLFB = imput[0:18]
        Z = "/"
        barvaZ = "#dddddd"
        #print(MLFB)
        #print(z)
        datum = imput[22:24]
        #print(datum)
        
        vc = imput[22:38]
        wc = len(vc)
        #print(wc)
        if wc == 13 :

            cislo = imput[22:26]+"-"+imput[26:30]+"-"+imput[30:32]+"-"+imput[32:35]
            #print(cislo)

        elif wc == 14:     

            cislo = imput[22:27]+"-"+imput[27:31]+"-"+imput[31:33]+"-"+imput[33:36]
            #print(cislo)
        

    #tvorba indexu 1ft6108a-8
    index = imput[0:6]+imput[15]+imput[7:9]
    print(index)

    #vyhledání indexu v tabulce
    for r in ws.rows:
        if r[0].value == index:
            odmer = r[1].value
            barvaO = "#dddddd"

    try:
        odmer
    except NameError:
        #print("Variable is not defined....!")
        odmer = "?"
        barvaO = "#ff0000"
        print(odmer)
    else:
        #print("Variable is defined.")
        print(odmer)
    
    #vyhledání data v tabulce
    for r in ws.rows:
        if wc == 13 :
            if r[2].value == datum:
                dat_v = r[3].value
                #print(dat_v)

        if wc == 14 :
            if r[4].value == datum:
                dat_v = r[5].value
                #print(dat_v)
    return(MLFB,cislo,odmer,Z,dat_v,barvaO,barvaZ)	

def vyroba2(cislo):
    wb = load_workbook(filename = 'data/data.xlsx')
    ws = wb['nic']
    cislo = cislo.replace("-","")
    cislo = cislo.replace(" ","")
    wc = len(cislo)
    datum = cislo[0:2]

    if not wc == 13|14:
        #print("nic")
        dat_v = "leden 2000"
    
    #vyhledání data v tabulce
    for r in ws.rows:
        if wc == 13 :
            if r[2].value == datum:
                dat_v = r[3].value
                #print(dat_v)

        if wc == 14 :
            if r[4].value == datum:
                dat_v = r[5].value
                #print(dat_v)
    return(dat_v)	
