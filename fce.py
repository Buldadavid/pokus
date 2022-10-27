from openpyxl import load_workbook
import csv
import inquirer
import os
    
#složka ve solici /diaglist    
folder = r'V:\Diag. l'

od = ""

while od == "":
    od = input("Začátek? ")

sub_folders = [name for name in os.listdir(folder) if os.path.isdir(os.path.join(folder, name))]
sub_folders = sorted(sub_folders)
#print(sub_folders)

filtered = []

for firma in sub_folders:
    if firma.startswith(od):
        filtered.append(firma)

#print(filtered)

questions = [inquirer.List("size",message="Firma?",choices=filtered, ),]

answers = inquirer.prompt(questions)

#print(answers["size"])
firma = answers["size"]

sf = ""
while sf == "":
    sf = input("SF? ")
    
akz = ""
while akz == "":
    akz = input("AKZ? ")
dbz = ""
while dbz == "":
    dbz = input("DBZ? ")
web = ""
while web == "":
    web = input("WB? ") 

#print(firma,akz,dbz,web)

datum = '=YEAR(TODAY())&"."&MONTH(TODAY())&"."&DAY(TODAY())'

header = ['QR','Firma','AKZ','DBZ','WB','MLFB','Z','cislo','Datum']
row = []

#nahadit diagl2.3.xlsm
wb = load_workbook(filename = 'Sešit1.xlsx')
ws = wb['List1']
#print(ws['A1'].value)

#nahadit adresou kde je uloženo na serveru
with open(r"Y:\WebAPP\data\data.csv") as file:
    reader = csv.DictReader(file)

    data = []

    for r in reader:
        data.append({"QR": r["QR"],"MLFB": r["MLFB"], "cislo": r["cislo"],"z": r["z"], "brzda": r["brzda"], "odmer": r["odmer"], "vaha": r["vaha"], "dat_v": r["dat_v"], "tep": r["tep"], "barvaO": r["barvaO"], "barvaZ": r["barvaZ"]})

QR = (data[0])["QR"]
MLFB = (data[0])["MLFB"]
cislo = (data[0])["cislo"]
Z = (data[0])["z"]
dat_v = (data[0])["dat_v"]
odmer = (data[0])["odmer"]
brzda = (data[0])["brzda"]
vaha = (data[0])["vaha"]
tep = (data[0])["tep"]
barvaO = (data[0])["barvaO"]
barvaZ = (data[0])["barvaZ"]

row.append(firma)
row.append(akz)
row.append(dbz)
row.append(web)
row.append(MLFB)
row.append(Z)
row.append(cislo)
row.append(datum)

with open('dtbz.csv', 'a') as f:
    writer = csv.writer(f)
    writer.writerow(row)

ws.insert_rows(1)
ws.insert_rows(1)
ws.insert_rows(1)
ws.insert_rows(1)
ws.insert_rows(1)

ws['C1'] = firma
ws['C2'] = akz
ws['C3'] = dbz
ws['C4'] = web

ws['E1'] = "Typ"
ws['E2'] = "Z"
ws['E3'] = "Výrobní číslo"
ws['E4'] = "Datum testu"

ws['F1'] = MLFB
ws['F2'] = Z
ws['F3'] = cislo
ws['F4'] = datum

ws['H4'] = sf

#nahadit adresou kam budu ukladat - složka jen na transfer po uložení smazat položky
wb.save("Sešit1.xlsx")
