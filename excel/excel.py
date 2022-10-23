from openpyxl import load_workbook
import csv
from datetime import *

today = date.today()

#print(today.strftime("%Y.%m.%d"))
datum = today.strftime("%Y.%m.%d")

header = ['QR','Firma','AKZ','DBZ','WB','MLFB','Z','cislo','Datum']
row = []

#nahadit diagl2.3.xlsm
wb = load_workbook(filename = 'zaklad.xlsx')
ws = wb['Sheet1']
#print(ws['A1'].value)

#nahadit adresou kde je uloženo na serveru
with open("/home/pi/Desktop/stitekweb-main/WebAPP/data/data.csv") as file:
    reader = csv.DictReader(file)

    data = []

    for r in reader:
        data.append({"QR": r["QR"],"MLFB": r["MLFB"], "cislo": r["cislo"],"z": r["z"], "brzda": r["brzda"], "odmer": r["odmer"], "vaha": r["vaha"], "dat_v": r["dat_v"], "tep": r["tep"], "barvaO": r["barvaO"], "barvaZ": r["barvaZ"]})

QR = (data[0])["QR"]
MLFB = (data[0])["MLFB"]
cislo = (data[0])["cislo"]
Z = (data[0])["z"]
dat_v = (data[0])["dat_v"]
odmer = (data[0])["odmer"]
brzda = (data[0])["brzda"]
vaha = (data[0])["vaha"]
tep = (data[0])["tep"]
barvaO = (data[0])["barvaO"]
barvaZ = (data[0])["barvaZ"]

firma = input("firma? ")
akz = input("akz? ")
dbz = input("dbz? ")
web = input("wb? ")

row.append(firma)
row.append(akz)
row.append(dbz)
row.append(web)
row.append(MLFB)
row.append(Z)
row.append(cislo)
row.append(datum)

with open('dtbz.csv', 'a') as f:
    writer = csv.writer(f)
    writer.writerow(row)

ws['B1'] = firma
ws['B2'] = akz
ws['B3'] = dbz
ws['B4'] = web

ws['D1'] = MLFB
ws['D2'] = Z
ws['D3'] = cislo
ws['D4'] = datum

#nahadit adresou kam budu ukladat - složka jen na transfer po uložení smazat položky
wb.save(f"tr/{firma}_{akz[0:6]+'-'+akz[6:8]+'_'+MLFB}.xlsx")